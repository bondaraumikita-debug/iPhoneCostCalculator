from openpyxl import Workbook
from openpyxl.styles import Font, numbers
import os

# Путь сохранения файла
file_path = r"D:\Python projects\P4Git\iPhone_17_Pro_Max_Calc.xlsx"
os.makedirs(os.path.dirname(file_path), exist_ok=True)

wb = Workbook()
ws = wb.active
ws.title = "Экономика"

# --- Параметры ---
ws["A1"] = "ПАРАМЕТРЫ"
ws["A1"].font = Font(bold=True)

params = [
    ("Цена за шт, $", 1400),
    ("Количество, шт", 50),
    ("Доставка, $", 300),
    ("Серт 024 (связь), $", 500),
    ("Серт 020, $", 300),
    ("Серт 037, $", 200),
    ("Утильсбор, %", 3),
    ("Таможня, %", 10),
    ("НДС, %", 20),
    ("Маржа, %", 3),
]

row = 2
for name, value in params:
    ws[f"A{row}"] = name
    ws[f"B{row}"] = value
    row += 1


# Заголовок блока расчётов
start_row = row + 1
ws[f"A{start_row}"] = "РАСЧЁТ"
ws[f"A{start_row}"].font = Font(bold=True)

headers = [
    "Закупка, $", "Утильсбор, $", "Доставка, $",
    "Серт. расходы, $", "Таможня, $",
    "СС без НДС, $", "НДС, $",
    "Маржа, $", "Итог цена парт, $",
    "Цена 1шт, $ без НДС", "Цена 1шт, $ с НДС"
]

col_row = start_row + 1
for col, h in enumerate(headers, start=1):
    cell = ws.cell(row=col_row, column=col, value=h)
    cell.font = Font(bold=True)

data_row = col_row + 1

# Формулы Excel
ws[f"A{data_row}"] = "=B2*B3"                                  # Закупка
ws[f"B{data_row}"] = f"=A{data_row}*(B8/100)"                  # Утильсбор
ws[f"C{data_row}"] = "=B4"                                     # Доставка
ws[f"D{data_row}"] = "=B5 + B6 + B7"                           # Сумма сертификатов
ws[f"E{data_row}"] = f"=A{data_row}*(B9/100)"                  # Таможня
ws[f"F{data_row}"] = f"=A{data_row}+B{data_row}+C{data_row}+D{data_row}+E{data_row}"  # СС без НДС
ws[f"G{data_row}"] = f"=F{data_row}*(B10/100)"                 # НДС
ws[f"H{data_row}"] = f"=F{data_row}*(B11/100)"                 # Маржа
ws[f"I{data_row}"] = f"=F{data_row}+G{data_row}+H{data_row}"   # Итог цена парт
ws[f"J{data_row}"] = f"=I{data_row}/B3"                        # Цена 1шт без НДС
ws[f"K{data_row}"] = f"=I{data_row}/B3"                        # Цена 1шт с НДС (включено в I)

# Формат валюты и ширина столбцов
for col in "ABCDEFGHIJK":
    ws[f"{col}{data_row}"].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
    ws.column_dimensions[col].width = 22

wb.save(file_path)

try:
    os.startfile(file_path)
except:
    pass

print("✔ Расчёт выполнен | 📁 Файл создан:", file_path)
